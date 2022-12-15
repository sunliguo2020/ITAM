import os

from django.conf import settings
from django.shortcuts import render, redirect, HttpResponse
from django.utils import timezone

from . import models
from app1.utils.form import ComputerModelForm, UserModelForm, DepModelForm
from app1.utils.pageination import Pagination

from openpyxl import load_workbook


# Create your views here.
def computer_list(request):
    # 构造搜索
    data_dict = {}
    search_data = request.GET.get('q', "")
    search_type = request.GET.get('search_type', '')

    if search_data:
        # 查询条件
        if search_type == 'seri':
            data_dict['serial_number__contains'] = search_data
            queryset = models.Computer.objects.filter(**data_dict)
        elif search_type == 'owner':
            # data_dict['owner__contains'] = search_data
            owner_obj = models.User.objects.filter(username__contains=search_data)
            # print(owner_obj)
            if owner_obj is not None:
                # 模糊查询
                queryset = models.Computer.objects.none()
                for item in owner_obj:
                    queryset = queryset | item.owners.all()
            else:  # 没有搜到使用者
                queryset = models.Computer.objects.none()

    else:
        queryset = models.Computer.objects.filter()

    obj = Pagination(request, queryset, page_size=10)
    context = {
        "queryset": obj.page_queryset,
        'page_string': obj.html()
    }
    return render(request, 'computer_list.html', context)


def computer_add(request):
    """
    电脑添加页面
    :param request:
    :return:
    """
    if request.method == "GET":
        form = ComputerModelForm()
        context = {
            "form": form
        }
        return render(request, 'computer_add.html', context)

    form = ComputerModelForm(data=request.POST, files=request.FILES)
    if form.is_valid():
        form.save()
        return redirect('/computer/list/')
    return render(request, 'computer_add.html', {'form': form})


def computer_edit(request, nid):
    """电脑编辑"""
    title = '电脑编辑'
    row_obj = models.Computer.objects.filter(id=nid).first()

    if request.method == "GET":
        form = ComputerModelForm(instance=row_obj)
        context = {
            'form': form,
            'title': title
        }
        return render(request, 'change.html', context)

    # 忘记添加files参数导致不能修改
    # form 中没有添加 enctype="multipart/form-data"
    form = ComputerModelForm(data=request.POST,files=request.FILES, instance=row_obj)
    if form.is_valid():
        form.save()
        return redirect('/computer/list/')
    context = {
        'form': form,
        'title': title
    }
    return render(request, 'change.html', context)


def computer_delete(request, nid):
    """
    删除电脑
    """
    # 删除文件
    file_path = models.Computer.objects.filter(id=nid).values('img')
    # print(file_path[0].get('img'))
    file_full_path = os.path.join(settings.MEDIA_ROOT, file_path[0].get('img'))
    if os.path.isfile(file_full_path):
        os.remove(file_full_path)
    """电脑记录删除"""
    models.Computer.objects.filter(id=nid).delete()

    return redirect('/computer/list/')


def user_list(request):
    # 构造搜索
    data_dict = {}
    search_data = request.GET.get('q', "")

    if search_data:
        # 查询条件
        data_dict['username__contains'] = search_data

    queryset = models.User.objects.filter(**data_dict)
    page_obj = Pagination(request, queryset)
    context = {
        'queryset': page_obj.page_queryset,
        'page_string': page_obj.html()
    }
    return render(request, 'user_list.html', context)


def user_add(request):
    """
    用户添加页面
    :param request:
    :return:
    """
    if request.method == "GET":
        form = UserModelForm()
        context = {
            "form": form
        }
        return render(request, 'user_add.html', context)
    form = UserModelForm(request.POST)
    if form.is_valid():
        form.save()

        return redirect('/user/list/')
    return render(request, 'user_add.html', {'form': form})


def user_edit(request, nid):
    # row_obj = models.User.objects.filter(id=nid).first()
    row_obj = models.User.objects.get(id=nid)
    if request.method == "GET":
        form = UserModelForm(instance=row_obj)
        context = {
            'form': form
        }
        return render(request, 'change.html', context)
    form = UserModelForm(data=request.POST, instance=row_obj)
    if form.is_valid():
        form.save()
        return redirect('/user/list/')

    context = {
        "form": form
    }
    return render(request, 'change.html', context)


def dep_list(request):
    queryset = models.Department.objects.all()
    context = {
        'queryset': queryset
    }

    return render(request, 'dep_list.html', context)


def dep_add(request):
    """
     部门添加页面
     :param request:
     :return:
     """
    if request.method == "GET":
        form = DepModelForm()
        context = {
            "form": form
        }
        return render(request, 'dep_add.html', context)
    form = DepModelForm(request.POST)
    if form.is_valid():
        form.save()
        return redirect('/dep/list/')
    return render(request, 'dep_add.html', {'form': form})


def user_delete(request, nid):
    models.User.objects.filter(id=nid).delete()

    return redirect('/user/list/')


def dep_delete(request, nid):
    models.Department.objects.filter(id=nid).delete()
    return redirect('/dep/list/')


def dep_edit(request, nid):
    row_obj = models.Department.objects.filter(id=nid).first()
    if request.method == "GET":
        form = DepModelForm(instance=row_obj)
        context = {
            "form": form
        }
        return render(request, 'change.html', context)

    form = DepModelForm(data=request.POST, instance=row_obj)
    context = {
        'form': form
    }
    if form.is_valid():
        form.save()

        return redirect('/dep/list/')
    return render(request, 'change.html', context)


def computer_multi(request):
    """
    批量上传文件导入到电脑表中
    :param request:
    :return:
    """
    # 是否上传空文件
    file_obj = request.FILES.get('excel', None)
    if not file_obj:
        return redirect('/dep/list/')

    # print(type(file_obj))
    # <class 'django.core.files.uploadedfile.TemporaryUploadedFile'>
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        # print(row[2].value)
        # brand, computer_type, serial_number, owner, production_date, mac_addr = [row[i].value for i in range(6)]
        # own_obj = models.User.objects.filter(username=owner).first()

        # 检查序列号是否已经存在
        if models.Computer.objects.filter(serial_number=row[2].value).exists():
            # print(row[2].value,'已经存在')
            continue

        data_dict = {
            'brand': row[0].value,
            'computer_type': row[1].value,
            'serial_number': row[2].value,
            'owner': models.User.objects.filter(username=row[3].value).first(),  # 查询用户表
            'production_date': row[4].value,
            'mac_addr': row[5].value,
        }
        # 判断生产日期的格式
        if not data_dict.get('production_date'):
            data_dict['production_date'] = timezone.now()

        models.Computer.objects.create(**data_dict)

    return redirect('/computer/list/')


def dep_multi(request):
    """
    批量上传文件导入到部门表
    :param request:
    :return:
    """
    file_obj = request.FILES.get('excel', None)
    if not file_obj:
        return redirect('/dep/list/')
    # 判断是否是excel文件
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        row_value = row[0].value
        if row[0].value and not models.Department.objects.filter(departname=row_value).exists():
            models.Department.objects.create(departname=row_value)
    return redirect('/dep/list/')


def user_multi(request):
    """
    批量上传文件导入到部门表
    :param request:
    :return:
    """
    file_obj = request.FILES.get('excel')
    if not file_obj:
        return redirect('/user/list/')
    # 判断是否是excel文件
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        row_value = row[0].value
        print(row_value)
        if row[0].value and not models.User.objects.filter(username=row_value).exists():
            models.User.objects.create(username=row_value, status=None)
    return redirect('/user/list/')
