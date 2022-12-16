# -*- coding: utf-8 -*-
"""
@author: sunliguo
@contact: QQ376440229
@Created on: 2022/12/16 8:31
"""
import os

from django.conf import settings
from django.shortcuts import render, redirect
from django.utils import timezone
from openpyxl import load_workbook

from app1 import models
from app1.utils.form import ComputerModelForm
from app1.utils.pageination import Pagination


def computer_list(request):
    # 构造搜索
    data_dict = {}
    search_data = request.GET.get('q', "")
    search_type = request.GET.get('search_type', '')

    if search_data:
        # 查询条件
        if search_type == 'seri':
            data_dict['serial_number__contains'] = search_data
            queryset = models.Computer.objects.filter(**data_dict)
        elif search_type == 'owner':
            # data_dict['owner__contains'] = search_data
            owner_obj = models.User.objects.filter(username__contains=search_data)
            # print(owner_obj)
            if owner_obj is not None:
                # 模糊查询
                queryset = models.Computer.objects.none()
                for item in owner_obj:
                    queryset = queryset | item.owners.all()
            else:  # 没有搜到使用者
                queryset = models.Computer.objects.none()

    else:
        queryset = models.Computer.objects.filter()

    obj = Pagination(request, queryset, page_size=10)
    context = {
        "queryset": obj.page_queryset,
        'page_string': obj.html()
    }
    return render(request, 'computer_list.html', context)


def computer_add(request):
    """
    电脑添加页面
    :param request:
    :return:
    """
    title = "新增电脑"
    if request.method == "GET":
        form = ComputerModelForm()
        context = {
            "form": form,
            'title': title
        }
        return render(request, 'add.html', context)

    form = ComputerModelForm(data=request.POST, files=request.FILES)

    if form.is_valid():
        form.save()
        return redirect('/computer/list/')
    else:
        print(form.cleaned_data)

    return render(request, 'computer_add.html', {'form': form})


def computer_edit(request, nid):
    """电脑编辑"""
    title = '电脑编辑'
    row_obj = models.Computer.objects.filter(id=nid).first()

    if request.method == "GET":
        form = ComputerModelForm(instance=row_obj)
        context = {
            'form': form,
            'title': title
        }
        return render(request, 'change.html', context)

    # 忘记添加files参数导致不能修改
    # form 中没有添加 enctype="multipart/form-data"
    form = ComputerModelForm(data=request.POST, files=request.FILES, instance=row_obj)
    if form.is_valid():
        form.save()
        return redirect('/computer/list/')
    context = {
        'form': form,
        'title': title
    }
    return render(request, 'change.html', context)


def computer_delete(request, nid):
    """
    删除电脑
    """
    # 删除文件
    file_path = models.Computer.objects.filter(id=nid).values('img').first()

    file_full_path = os.path.join(settings.MEDIA_ROOT, file_path.get('img'))
    # 不删除默认图片
    if os.path.isfile(file_full_path) and file_path.get('img') != 'computer/none.jpg':
        os.remove(file_full_path)

    """电脑记录删除"""
    models.Computer.objects.filter(id=nid).delete()

    return redirect('/computer/list/')


def computer_multi(request):
    """
    批量上传文件导入到电脑表中
    :param request:
    :return:
    """
    # 是否上传空文件
    file_obj = request.FILES.get('excel', None)
    if not file_obj:
        return redirect('/dep/list/')

    # print(type(file_obj))
    # <class 'django.core.files.uploadedfile.TemporaryUploadedFile'>
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        # print(row[2].value)
        # brand, computer_type, serial_number, owner, production_date, mac_addr = [row[i].value for i in range(6)]
        # own_obj = models.User.objects.filter(username=owner).first()

        # 检查序列号是否已经存在
        if models.Computer.objects.filter(serial_number=row[2].value).exists():
            # print(row[2].value,'已经存在')
            continue

        data_dict = {
            'brand': row[0].value,
            'computer_type': row[1].value,
            'serial_number': row[2].value,
            'owner': models.User.objects.filter(username=row[3].value).first(),  # 查询用户表
            'production_date': row[4].value,
            'mac_addr': row[5].value,
        }
        # 判断生产日期的格式
        if not data_dict.get('production_date'):
            data_dict['production_date'] = timezone.now()

        models.Computer.objects.create(**data_dict)

    return redirect('/computer/list/')
