# -*- coding: utf-8 -*-
"""
@author: sunliguo
@contact: QQ376440229
@Created on: 2022/12/16 8:36
"""

from django.shortcuts import render, redirect

from app1 import models
from app1.utils.form import DepModelForm
from app1.utils.pageination import Pagination

from openpyxl import load_workbook


def dep_list(request):
    queryset = models.Department.objects.all()
    page_queryset = Pagination(request,queryset)
    context = {
        'queryset': page_queryset.page_queryset,
        'page_string':page_queryset.html()
    }

    return render(request, 'dep_list.html', context)


def dep_add(request):
    """
     部门添加页面
     :param request:
     :return:
     """
    title = '添加部门'
    if request.method == "GET":
        form = DepModelForm()
        context = {
            "form": form,
            'title':title
        }
        return render(request, 'add.html', context)
    form = DepModelForm(request.POST)
    if form.is_valid():
        form.save()
        return redirect('/dep/list/')
    return render(request, 'add.html', {'form': form})


def dep_delete(request, nid):
    models.Department.objects.filter(id=nid).delete()
    return redirect('/dep/list/')


def dep_edit(request, nid):
    row_obj = models.Department.objects.filter(id=nid).first()
    if request.method == "GET":
        form = DepModelForm(instance=row_obj)
        context = {
            "form": form
        }
        return render(request, 'change.html', context)

    form = DepModelForm(data=request.POST, instance=row_obj)
    context = {
        'form': form
    }
    if form.is_valid():
        form.save()

        return redirect('/dep/list/')
    return render(request, 'change.html', context)


def dep_multi(request):
    """
    批量上传文件导入到部门表
    :param request:
    :return:
    """
    file_obj = request.FILES.get('excel', None)
    if not file_obj:
        return redirect('/dep/list/')
    # 判断是否是excel文件
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        row_value = row[0].value
        if row[0].value and not models.Department.objects.filter(departname=row_value).exists():
            models.Department.objects.create(departname=row_value)
    return redirect('/dep/list/')
