# -*- coding: utf-8 -*-
"""
@author: sunliguo
@contact: QQ376440229
@Created on: 2022-12-16 18:32
"""
import csv

import pytz
from django.shortcuts import render, redirect, HttpResponse
from openpyxl import load_workbook

from app1 import models
from app1.utils.form import IpModelForm
from app1.utils.pageination import Pagination


def ip_list(request):
    queryset = models.IpAddr.objects.all()
    page_object = Pagination(request, queryset)
    context = {
        'queryset': page_object.page_queryset,
        'page_string': page_object.html()
    }
    return render(request, 'ip_list.html', context)


def ip_add(request):
    if request.method == 'GET':
        title = '新增Ip地址'
        form = IpModelForm()
        context = {
            'form': form
        }
        return render(request, 'add.html', context)
    form = IpModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect('/ip/list/')
    context = {
        'form': form
    }
    return render(request, 'add.html', context)


def ip_edit(request, nid):
    obj = models.IpAddr.objects.filter(id=nid).first()
    if request.method == 'GET':
        form = IpModelForm(instance=obj)
        context = {
            'form': form
        }
        return render(request, 'change.html', context)
    form = IpModelForm(data=request.POST, instance=obj)
    if form.is_valid():
        form.save()
        return redirect('/ip/list/')
    context = {
        'form': form
    }
    return render(request, 'change.html', context)


def ip_delete(request, nid):
    models.IpAddr.objects.filter(id__gt=100).delete()
    return redirect('/ip/list/')


def ip_multi(request):
    """
    ip arp 地址对应关系导入
    :param request:
    :return:
    """
    # 是否上传空文件
    file_obj = request.FILES.get('csv_file', None)
    if not file_obj:
        return redirect('/ip/list/')

    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=1):
        data_dict = {
            'ip_addr': row[0].value,
            'mac_addr': row[1].value,
            'interface': row[2].value,
            # 解决时区问题
            'cap_datetime': row[3].value.replace(tzinfo=pytz.timezone('UTC'))
        }

        # 检查前三项是否已经有值，如果日期更新，则更新日期
        queryset = models.IpAddr.objects.filter(ip_addr=data_dict['ip_addr'],
                                                mac_addr=data_dict['mac_addr'],
                                                interface=data_dict['interface'])

        if queryset.exists():
            print('已经有数据', queryset)
            for query_row in queryset:
                print('cap_datetime', data_dict['cap_datetime'])
                print('query_row.cap_datetime', query_row.cap_datetime)
                if data_dict['cap_datetime'] > query_row.cap_datetime:
                    print('更新时间',type(query_row))
                    query_row.cap_datetime = data_dict['cap_datetime']
                    query_row.save()
        else:
            print('新增数据')
            models.IpAddr.objects.create(**data_dict)
        # 检查是否有重复值
        # if not models.IpAddr.objects.filter(**data_dict).exists():
        #     models.IpAddr.objects.create(**data_dict)
        # else:
        #     print('有重复值')
        # if models.IpAddr.objects.get_or_create(**data_dict):
        #     print('有重复值')
        # else:
        #     print('没有重复值')

    return redirect('/ip/list/')
