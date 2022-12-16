# -*- coding: utf-8 -*-
"""
@author: sunliguo
@contact: QQ376440229
@Created on: 2022/12/16 8:34
"""
from app1 import models
from django.shortcuts import render, redirect, HttpResponse

from app1.utils.form import  UserModelForm, DepModelForm
from app1.utils.pageination import Pagination

from openpyxl import load_workbook

def user_list(request):
    # 构造搜索
    data_dict = {}
    search_data = request.GET.get('q', "")

    if search_data:
        # 查询条件
        data_dict['username__contains'] = search_data

    queryset = models.User.objects.filter(**data_dict)
    page_obj = Pagination(request, queryset)
    context = {
        'queryset': page_obj.page_queryset,
        'page_string': page_obj.html()
    }
    return render(request, 'user_list.html', context)


def user_add(request):
    """
    用户添加页面
    :param request:
    :return:
    """
    if request.method == "GET":
        form = UserModelForm()
        context = {
            "form": form
        }
        return render(request, 'user_add.html', context)
    form = UserModelForm(request.POST)
    if form.is_valid():
        form.save()

        return redirect('/user/list/')
    return render(request, 'user_add.html', {'form': form})


def user_edit(request, nid):
    # row_obj = models.User.objects.filter(id=nid).first()
    row_obj = models.User.objects.get(id=nid)
    if request.method == "GET":
        form = UserModelForm(instance=row_obj)
        context = {
            'form': form
        }
        return render(request, 'change.html', context)
    form = UserModelForm(data=request.POST, instance=row_obj)
    if form.is_valid():
        form.save()
        return redirect('/user/list/')

    context = {
        "form": form
    }
    return render(request, 'change.html', context)


def user_delete(request, nid):
    models.User.objects.filter(id=nid).delete()

    return redirect('/user/list/')


def user_multi(request):
    """
    批量上传文件导入到部门表
    :param request:
    :return:
    """
    file_obj = request.FILES.get('excel')
    if not file_obj:
        return redirect('/user/list/')
    # 判断是否是excel文件
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        row_value = row[0].value
        print(row_value)
        if row[0].value and not models.User.objects.filter(username=row_value).exists():
            models.User.objects.create(username=row_value, status=None)
    return redirect('/user/list/')
